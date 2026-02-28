"""Extract StereoSet examples and add to test plan Excel.

Creates pairs:
  - context + stereotype → Biased
  - context + anti-stereotype → Neutral
"""
import sys
import io
import json
import pandas as pd
from openpyxl import load_workbook

# Fix encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load StereoSet data
stereoset_path = 'attention_app/bias/stereoset/results/stereoset_precomputed_gusnet.json'
with open(stereoset_path, 'r', encoding='utf-8') as f:
    data = json.load(f)

examples = data['examples']
print(f"Total StereoSet examples: {len(examples)}")

# Read current Excel
excel_path = 'Plano_Testes_AttentionBias_AnaPinto.xlsx'
df_textos = pd.read_excel(excel_path, sheet_name='Textos')
print(f"Current Textos sheet: {len(df_textos)} rows")

# Find next available ID
current_ids = df_textos['ID'].tolist()
biased_ids = [id for id in current_ids if id.startswith('B')]
neutral_ids = [id for id in current_ids if id.startswith('N')]

# Extract last numbers
last_b = max([int(id[1:]) for id in biased_ids if id[1:].isdigit()], default=0)
last_n = max([int(id[1:]) for id in neutral_ids if id[1:].isdigit()], default=0)

print(f"Last biased ID: B{last_b}, Last neutral ID: N{last_n}")

# Add all available examples
n_pairs = len(examples)
print(f"\nAdding all {n_pairs} StereoSet example pairs ({n_pairs} biased + {n_pairs} neutral)")

# Build new rows
new_rows = []
for i, ex in enumerate(examples[:n_pairs]):
    # Biased: context + stereotype
    biased_text = f"{ex['context']} {ex['stereo_sentence']}"
    biased_id = f"B{last_b + i + 1}"

    new_rows.append({
        'ID': biased_id,
        'Etiqueta esperada': 'Biased',
        'Tema': f"StereoSet-{ex['category']}",
        'Texto a usar': biased_text,
        'Spans': '',
        'Porque é bias': f"Stereotype about {ex['target']}",
        'Categorias esperadas': 'STEREO',
        'Sinais esperados': f"Target: {ex['target']}",
        'Notas': f"StereoSet context+stereo, category={ex['category']}",
    })

    # Neutral: context + anti-stereotype
    neutral_text = f"{ex['context']} {ex['anti_sentence']}"
    neutral_id = f"N{last_n + i + 1}"

    new_rows.append({
        'ID': neutral_id,
        'Etiqueta esperada': 'Neutral',
        'Tema': f"StereoSet-{ex['category']}",
        'Texto a usar': neutral_text,
        'Spans': '',
        'Porque é bias': '',
        'Categorias esperadas': '',
        'Sinais esperados': '',
        'Notas': f"StereoSet context+anti-stereo, category={ex['category']}",
    })

df_new = pd.DataFrame(new_rows)

# Add empty columns to match Excel structure
for col in df_textos.columns:
    if col not in df_new.columns:
        df_new[col] = ''

# Reorder columns
df_new = df_new[df_textos.columns]

# Concatenate
df_combined = pd.concat([df_textos, df_new], ignore_index=True)

print(f"\nNew total: {len(df_combined)} rows")
print(f"Added {len(new_rows)} sentences ({n_pairs} biased + {n_pairs} neutral)")

# Preview
print("\nPreview of first 4 new sentences:")
for i in range(min(4, len(new_rows))):
    row = new_rows[i]
    print(f"\n{row['ID']} ({row['Etiqueta esperada']}):")
    print(f"  {row['Texto a usar'][:80]}...")

# Write to Excel
wb = load_workbook(excel_path)
if 'Textos' in wb.sheetnames:
    del wb['Textos']
ws = wb.create_sheet('Textos', 0)

# Header
for col_idx, col_name in enumerate(df_combined.columns, start=1):
    ws.cell(row=1, column=col_idx, value=col_name)

# Data
for row_idx, (_, row) in enumerate(df_combined.iterrows(), start=2):
    for col_idx, val in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=val)

wb.save(excel_path)
print(f"\nSaved to {excel_path}")
print(f"  Total sentences: {len(df_combined)}")
print(f"  Biased: {len(df_combined[df_combined['ID'].str.startswith('B')])}")
print(f"  Neutral: {len(df_combined[df_combined['ID'].str.startswith('N')])}")
