"""Automated Test-Plan Analysis — Plano de Testes AttentionBias.

Reads test sentences from the Excel test plan, runs each one through
both BERT and GPT-2 pipelines, and produces an Excel file mirroring
the "Registo" sheet — plus extra faithfulness columns:

  Mean BAR_i, Mean IG rho_i, Significant heads (count/total),
  Max Ablation Delta_i, Mean Jaccard, Mean RBO (p=0.9),
  Mean Spearman rho, rho(Perturb,IG), Mean rho(Perturb,Attn),
  Max Perturbation Impact

Usage:
    py run_test_plan_analysis.py                            # 16 built-in test sentences, both models
    py run_test_plan_analysis.py --model bert               # BERT only
    py run_test_plan_analysis.py --model gpt2               # GPT-2 only
    py run_test_plan_analysis.py --input frases.csv         # custom sentences from CSV
    py run_test_plan_analysis.py --input frases.xlsx        # custom sentences from Excel
    py run_test_plan_analysis.py --out results.xlsx         # custom output path
    py run_test_plan_analysis.py --n-steps 20               # fewer IG steps (faster)
    py run_test_plan_analysis.py --template                 # generate template CSV

Input file format (CSV or Excel):
    id,expected,text
    B1,biased,"For a customer support role, women are usually..."
    N1,neutral,"For a customer support role, we look for..."

    Columns: id (required), text (required), expected (optional: biased/neutral)
"""

import sys, os, argparse, time, warnings, subprocess
from datetime import datetime

# Fix Windows console encoding for GPT-2 tokens (Ġ etc.)
if sys.stdout.encoding and sys.stdout.encoding.lower().startswith("cp"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import numpy as np
import pandas as pd
import torch
import joblib

#  Project imports ─
sys.path.insert(0, os.path.dirname(__file__))

from attention_app.models import ModelManager
from attention_app.bias.attention_bias import AttentionBiasAnalyzer
from attention_app.bias.gusnet_detector import GusNetDetector
from attention_app.bias.head_ablation import batch_ablate_top_heads
from attention_app.bias.integrated_gradients import (
    batch_compute_ig_correlation,
    batch_compute_perturbation,
    batch_compute_lrp,
)
from attention_app.bias.feature_extraction_notebooks import extract_features_for_sentence

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)


#  Test sentences 

TEST_SENTENCES = [
    ("B1", "biased",
     "For a customer support role, women are usually a safer choice because "
     "they tend to be more patient and caring."),
    ("B2", "biased",
     "If the team includes older employees, it\u2019s better to avoid frequent "
     "software changes because they often take longer to adapt."),
    ("B3", "biased",
     "The applicant has a strong foreign accent, so clients might not take "
     "him seriously."),
    ("N1", "neutral",
     "For a customer support role, we look for clear communication, patience "
     "under pressure, and the ability to handle multiple requests."),
    ("N2", "neutral",
     "Before rolling out frequent software changes, provide training and time "
     "to adapt, especially for people who do not use the tool daily."),
    ("N3", "neutral",
     "The applicant has a strong accent, but his speech is clear; we\u2019ll "
     "evaluate communication skills in the interview."),
    ("B4", "biased",  "The problem with CEOs is that they are greedy."),
    ("B5", "biased",  "Men are inherently promiscuous."),
    ("B6", "biased",
     "Trump, who has faced criticism over his handling of the pandemic, said "
     "China must be held accountable for failing to contain the virus."),
    ("B7", "biased",  "Every single CEOs is fake."),
    ("B8", "biased",  "Cyclists are naturally narcissistic."),
    ("N4", "neutral",
     "Does this research reveal any trends regarding the prevalence of mental "
     "health conditions among teenagers?"),
    ("N5", "neutral",
     "It is therefore no surprise that Ocasio-Cortez has embraced the term "
     "\u201cDemocratic Socialist.\u201d"),
    ("N6", "neutral",
     "Thermal energy transfers through conduction, convection, and radiation."),
    ("N7", "neutral",
     "Caffeine can increase alertness and energy levels."),
    ("N8", "neutral",
     "What long-term effects might result from using hormone treatments aimed "
     "at increasing height during adolescence?"),
]


#  Input loading ─

def _load_sentences_from_file(path):
    """Load sentences from a CSV or Excel file.

    Expected columns: id, text, expected (optional).
    Also accepts: ID, Texto, Etiqueta (Portuguese variants).
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        # Try "Textos" sheet first (test plan layout), fallback to first sheet
        try:
            df = pd.read_excel(path, sheet_name="Textos")
        except ValueError:
            df = pd.read_excel(path)
    else:
        # Try common separators
        for sep in [",", ";", "\t"]:
            df = pd.read_csv(path, sep=sep, encoding="utf-8-sig")
            if len(df.columns) >= 2:
                break

    # Normalise column names
    col_map = {}
    for col in df.columns:
        cl = col.strip().lower()
        if cl in ("id", "id do texto", "text_id"):
            col_map[col] = "id"
        elif cl in ("text", "texto", "texto colado", "sentence", "frase",
                     "texto a usar"):
            col_map[col] = "text"
        elif cl in ("expected", "etiqueta", "etiqueta esperada", "label",
                     "esperado"):
            col_map[col] = "expected"
    df = df.rename(columns=col_map)

    if "text" not in df.columns:
        raise ValueError(
            f"Input file must have a 'text' column. Found: {list(df.columns)}"
        )
    if "id" not in df.columns:
        df["id"] = [f"S{i+1}" for i in range(len(df))]
    if "expected" not in df.columns:
        df["expected"] = ""

    # Normalise expected values
    def _norm_expected(val):
        val = str(val).strip().lower()
        if val in ("biased", "com bias", "bias", "1", "true"):
            return "biased"
        if val in ("neutral", "sem bias", "neutro", "0", "false"):
            return "neutral"
        return ""  # unknown
    df["expected"] = df["expected"].apply(_norm_expected)

    # Drop empty rows
    df = df.dropna(subset=["text"])
    df = df[df["text"].str.strip().astype(bool)]

    sentences = [
        (str(row["id"]).strip(), row["expected"], str(row["text"]).strip())
        for _, row in df.iterrows()
    ]
    print(f"Loaded {len(sentences)} sentences from {path}")
    return sentences


def _generate_template(path="frases_template.csv"):
    """Write a template CSV file with example sentences."""
    lines = [
        "id,expected,text",
        'B_ex1,biased,"Women are too emotional to be effective leaders."',
        'B_ex2,biased,"Old people can\'t learn new technology."',
        'N_ex1,neutral,"Effective leadership requires strong communication skills."',
        'N_ex2,neutral,"Learning new technology takes practice and patience."',
        'X_ex1,,"This sentence has no expected label — will still be analysed."',
    ]
    with open(path, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(lines) + "\n")
    print(f"Template written to: {path}")
    print("Edit the file, then run:")
    print(f"  py run_test_plan_analysis.py --input {path}")


# Model mappings 

GUSNET_KEY = {"bert": "gusnet-bert", "gpt2": "gusnet-gpt2"}
ENCODER_NAME = {"bert": "bert-base-uncased", "gpt2": "gpt2"}
CLASSIFIER_PATH = {
    "bert": os.path.join(os.path.dirname(__file__),
                         "attention_app", "bias", "saved_models",
                         "bert_bias_classifier.joblib"),
    "gpt2": os.path.join(os.path.dirname(__file__),
                         "attention_app", "bias", "saved_models",
                         "gpt2_bias_classifier.joblib"),
}


# Helpers 

def _get_git_commit():
    """Return the short commit hash of the current repo."""
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=os.path.dirname(__file__),
            stderr=subprocess.DEVNULL,
        ).decode().strip()
    except Exception:
        return "unknown"


def _format_top_heads(metrics, n=4):
    """Return e.g. 'L0-H1  L0-H8  L0-H5  L0-H0'."""
    top = sorted(metrics, key=lambda m: m.bias_attention_ratio, reverse=True)[:n]
    return "  ".join(f"L{m.layer}-H{m.head}" for m in top)


def _gusnet_weighted_score(summary):
    """Compute the weighted composite score matching the dashboard.

    Mirrors ``create_bias_criteria_html`` in visualizations.py (lines 1130-1156).

    Components (each normalised to [0, 1]):
        Token density  (bias_percentage / 100)            x 0.30
        Generalization (gen_count / total * 5, capped 1)  x 0.20
        Unfair language(unfair_count / total * 5, cap 1)  x 0.25
        Stereotypes    (stereo_count / total * 5, cap 1)  x 0.25

    Returns (score, level_str).
    """
    weights = {"pct": 0.30, "gen": 0.20, "unfair": 0.25, "stereo": 0.25}

    total = max(summary.get("total_tokens", 1), 1)
    pct = summary.get("bias_percentage", 0) / 100
    gen_r = summary.get("generalization_count", 0) / total
    unfair_r = summary.get("unfairness_count", 0) / total
    stereo_r = summary.get("stereotype_count", 0) / total

    c_pct = pct
    c_gen = min(gen_r * 5, 1.0)
    c_unfair = min(unfair_r * 5, 1.0)
    c_stereo = min(stereo_r * 5, 1.0)

    score = (
        weights["pct"] * c_pct
        + weights["gen"] * c_gen
        + weights["unfair"] * c_unfair
        + weights["stereo"] * c_stereo
    )
    score = min(score, 1.0)

    if score < 0.15:
        level = "Low"
    elif score < 0.40:
        level = "Moderate"
    else:
        level = "High"

    return score, level


def _notebook_classifier_predict(text, model_key):
    """Extract features and run the saved sklearn pipeline (notebook classifier).

    Returns (probability, label_str).
    """
    clf_path = CLASSIFIER_PATH[model_key]
    if not os.path.exists(clf_path):
        return None, None

    pipeline = joblib.load(clf_path)
    feature_names = list(pipeline.feature_names_in_)

    # Extract features using the same function used in the notebooks
    raw_features = extract_features_for_sentence(
        text, ENCODER_NAME[model_key], ModelManager
    )

    # Build feature vector in the correct column order
    feature_vector = np.array(
        [raw_features.get(name, 0.0) for name in feature_names],
        dtype=np.float64,
    ).reshape(1, -1)

    # Replace NaN / inf
    feature_vector = np.nan_to_num(feature_vector, nan=0.0, posinf=0.0, neginf=0.0)

    prob = float(pipeline.predict_proba(feature_vector)[0, 1])  # P(biased)
    label = "BIASED" if prob >= 0.5 else "NEUTRAL"
    return prob, label


#  Main analysis function 

def analyze_text(
    text_id, text, expected, model_key, commit_hash,
    n_steps=30, bar_threshold=1.5, sig_alpha=0.05,
):
    """Run the full analysis pipeline for one (text, model) pair."""
    is_gpt2 = model_key == "gpt2"
    encoder_name = ENCODER_NAME[model_key]
    gusnet_key = GUSNET_KEY[model_key]

    # 1. Load encoder
    tokenizer, encoder, mlm = ModelManager.get_model(encoder_name)
    device = next(encoder.parameters()).device

    # 2. Forward pass -> attention weights
    inputs = tokenizer(text, return_tensors="pt", truncation=True, max_length=512)
    inputs_dev = {k: v.to(device) for k, v in inputs.items()}
    with torch.no_grad():
        outputs = encoder(**inputs_dev, output_attentions=True)
    attentions = list(outputs.attentions)
    tokens = tokenizer.convert_ids_to_tokens(inputs_dev["input_ids"][0])

    num_layers = len(attentions)
    num_heads = attentions[0].shape[1]
    total_heads = num_layers * num_heads

    # 3. GUS-Net token-level bias detection
    gus = GusNetDetector(model_key=gusnet_key, use_optimized=True)
    token_labels = gus.detect_bias(text)
    biased_indices = [t["index"] for t in token_labels if t["is_biased"]]

    gus_summary = gus.get_bias_summary(token_labels)
    gus_score, gus_level = _gusnet_weighted_score(gus_summary)

    # 4. Notebook classifier
    nb_prob, nb_label = _notebook_classifier_predict(text, model_key)

    # 5. Attention x Bias (BAR / BSR)
    analyzer = AttentionBiasAnalyzer()
    attention_metrics = analyzer.analyze_attention_to_bias(
        attentions, biased_indices, tokens, bar_threshold=bar_threshold,
    )

    if attention_metrics:
        mean_bar = float(np.mean([m.bias_attention_ratio for m in attention_metrics]))
        max_bar = float(max(m.bias_attention_ratio for m in attention_metrics))
        n_sig_bar = sum(1 for m in attention_metrics if m.specialized_for_bias)
        top_heads_str = _format_top_heads(attention_metrics)
    else:
        mean_bar = max_bar = 0.0
        n_sig_bar = 0
        top_heads_str = "None"

    # 6. Integrated Gradients + Top-K Overlap (Jaccard, RBO)
    ig_bundle = batch_compute_ig_correlation(
        encoder, tokenizer, text, attentions,
        attention_metrics, is_gpt2, n_steps=n_steps,
    )
    ig_corrs = ig_bundle.correlations
    mean_ig_rho = float(np.mean([c.spearman_rho for c in ig_corrs])) if ig_corrs else 0.0
    n_ig_sig = sum(1 for c in ig_corrs if c.spearman_pvalue < sig_alpha)

    topk = ig_bundle.topk_overlaps or []
    mean_jaccard = float(np.mean([t.jaccard for t in topk])) if topk else 0.0
    mean_rbo = float(np.mean([t.rank_biased_overlap for t in topk])) if topk else 0.0

    # 7. Head Ablation
    if attention_metrics:
        top_abl = sorted(
            attention_metrics, key=lambda m: m.bias_attention_ratio, reverse=True
        )[:20]
        abl_results = batch_ablate_top_heads(
            encoder, mlm, tokenizer, text, top_abl, is_gpt2,
        )
        max_abl = float(max(r.representation_impact for r in abl_results)) if abl_results else 0.0
    else:
        max_abl = 0.0

    # 8. Perturbation Analysis
    perturb = batch_compute_perturbation(
        encoder, tokenizer, text, is_gpt2,
        ig_bundle.token_attributions, attentions,
    )
    rho_p_ig = perturb.perturb_vs_ig_spearman
    mean_rho_p_attn = (
        float(np.mean([r[2] for r in perturb.perturb_vs_attn_spearman]))
        if perturb.perturb_vs_attn_spearman else 0.0
    )
    max_perturb = (
        float(max(r.importance for r in perturb.token_results))
        if perturb.token_results else 0.0
    )

    # 8b. LRP Analysis
    try:
        lrp_bundle = batch_compute_lrp(
            encoder, tokenizer, text, is_gpt2,
            ig_bundle.token_attributions, attentions,
            attention_metrics,
        )
        rho_lrp_ig = float(lrp_bundle.lrp_vs_ig_spearman)
        mean_rho_lrp_attn = (
            float(np.mean([r[2] for r in lrp_bundle.correlations]))
            if lrp_bundle.correlations else 0.0
        )
    except Exception:
        rho_lrp_ig = 0.0
        mean_rho_lrp_attn = 0.0

    # 9. Biased spans and categories
    # Clean GPT-2 tokenizer artifacts (Ġ = space prefix)
    def clean_token(tok):
        return tok.replace("Ġ", " ").strip()

    biased_spans = ", ".join(
        clean_token(t["token"]) for t in token_labels if t["is_biased"]
    ) or "None"
    categories_str = "; ".join(
        f"{clean_token(t['token'])}-{','.join(t['bias_types'])}"
        for t in token_labels if t["is_biased"]
    ) or "None"

    # 10. Auto-evaluate pass/fail
    notes = []
    passed = True

    if expected:  # only evaluate if expected label is known
        is_bias_text = expected == "biased"

        if nb_label is not None:
            nb_correct = (nb_label == "BIASED") == is_bias_text
            if not nb_correct:
                notes.append(f"Notebook classifier: {nb_label} (expected {'BIASED' if is_bias_text else 'NEUTRAL'})")
                passed = False

        gus_correct = (gus_level != "Low") == is_bias_text
        if not gus_correct:
            if is_bias_text and gus_level == "Low":
                notes.append("GUS-Net bias too low for biased text")
                passed = False
            elif not is_bias_text and gus_level != "Low":
                notes.append(f"GUS-Net false positive ({gus_level})")
    else:
        passed = None  # no expected label → can't judge

    model_display = "GPT-2" if is_gpt2 else "BERT"

    return {
        #  Original "Registo" columns 
        "Data": datetime.now().strftime("%Y-%m-%d"),
        "Testador": "Auto",
        "Versao / commit": commit_hash,
        "Modelo": model_display,
        "ID do texto": text_id,
        "Texto colado": text,
        "Prob. classificador Notebook": round(nb_prob, 4) if nb_prob is not None else "",
        "Etiqueta classificador Notebook": nb_label or "",
        "Prob. classificador GUS-NET": round(gus_score, 4),
        "Etiqueta classificador GUS-NET": gus_level,
        "Bias Level (token-level)": gus_level,
        "Spans detectados": biased_spans,
        "Categorias (GEN/UNFAIR/STEREO)": categories_str,
        "Top heads (Attention-Bias)": top_heads_str,
        "Max bias ratio": round(max_bar, 3),
        "Notas / observacoes": "; ".join(notes) if notes else "",
        "Passou? (S/N)": "",
        # faithfulness metrics 
        "Mean Jaccard": round(mean_jaccard, 4),
        "Mean Spearman rho": round(mean_ig_rho, 4),
        "Mean RBO (p=0.9)": round(mean_rbo, 4),
        "rho(LRP, IG)": round(rho_lrp_ig, 4),
        "Mean rho(LRP, Attn)": round(mean_rho_lrp_attn, 4),
        "rho(Perturb, IG)": round(rho_p_ig, 4),
        "Mean rho(Perturb, Attn)": round(mean_rho_p_attn, 4),
        "Max Perturbation Impact": round(max_perturb, 6),
        "Mean BAR": round(mean_bar, 4),
        "Mean IG rho": round(mean_ig_rho, 4),
        "Max Ablation Delta": round(max_abl, 6),
    }


#  Excel output 

def _save_excel(df, path, preserve_sheets=False):
    """Save DataFrame to a styled Excel workbook.

    If *preserve_sheets* is True and *path* already exists, only the
    "Registo" sheet is replaced/created — all other sheets are kept.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if preserve_sheets and os.path.exists(path):
        wb = load_workbook(path)
        # Remove old Registo sheet if it exists
        if "Registo" in wb.sheetnames:
            del wb["Registo"]
        ws = wb.create_sheet("Registo")
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Registo"

    # Write header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    #  Styling 
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        bottom=Side(style="thin", color="B4C6E7"),
    )

    original_cols = 17  # columns up to Passou
    new_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = new_fill if col_idx > original_cols else header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    passou_col = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value and "Passou" in str(cell.value):
            passou_col = col_idx
            break

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

        if passou_col:
            val = ws.cell(row=row_idx, column=passou_col).value
            fill = green_fill if val == "S" else red_fill if val == "N" else None
            if fill:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    col_widths = {
        "Data": 12, "Testador": 8, "Versao / commit": 10, "Modelo": 8,
        "ID do texto": 10, "Texto colado": 40,
        "Prob. classificador Notebook": 14, "Etiqueta classificador Notebook": 12,
        "Prob. classificador GUS-NET": 14, "Etiqueta classificador GUS-NET": 12,
        "Bias Level (token-level)": 10,
        "Spans detectados": 30, "Categorias (GEN/UNFAIR/STEREO)": 30,
        "Top heads (Attention-Bias)": 25, "Max bias ratio": 10,
        "Notas / observacoes": 30, "Passou? (S/N)": 8,
    }
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_val = ws.cell(row=1, column=col_idx).value or ""
        width = col_widths.get(header_val, 14)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


# CLI entry-point

def main():
    parser = argparse.ArgumentParser(
        description="Automated Test-Plan Bias Analysis",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  py run_test_plan_analysis.py                         # built-in 16 test sentences
  py run_test_plan_analysis.py --input frases.csv      # your own sentences
  py run_test_plan_analysis.py --template              # generate template CSV
  py run_test_plan_analysis.py --input frases.csv --model bert --n-steps 50
""",
    )
    parser.add_argument(
        "--input", type=str, default=None,
        help="Path to CSV or Excel file with sentences (columns: id, text, expected)",
    )
    parser.add_argument(
        "--template", action="store_true",
        help="Generate a template CSV file and exit",
    )
    parser.add_argument(
        "--model", choices=["bert", "gpt2", "both"], default="both",
        help="Which model(s) to run (default: both)",
    )
    parser.add_argument(
        "--out", type=str, default="test_plan_results.xlsx",
        help="Output file path (default: test_plan_results.xlsx)",
    )
    parser.add_argument(
        "--n-steps", type=int, default=30,
        help="IG interpolation steps (default: 30)",
    )
    parser.add_argument(
        "--bar-threshold", type=float, default=1.5,
        help="BAR threshold for significant heads (default: 1.5)",
    )
    parser.add_argument(
        "--sig-alpha", type=float, default=0.05,
        help="Significance level for IG p-value (default: 0.05)",
    )
    parser.add_argument(
        "--old-runs", type=str, default=None,
        help="Path to Excel file with old executions to preserve (reads Registo sheet)",
    )
    args = parser.parse_args()

    # --template: generate example CSV and exit
    if args.template:
        _generate_template()
        return

    # Load sentences
    if args.input:
        sentences = _load_sentences_from_file(args.input)
    else:
        sentences = TEST_SENTENCES

    models = ["bert", "gpt2"] if args.model == "both" else [args.model]
    commit_hash = _get_git_commit()

    rows = []
    total = len(sentences) * len(models)
    done = 0

    for model_key in models:
        print(f"\n{'='*60}")
        print(f"  Model: {model_key.upper()}  |  commit: {commit_hash}")
        print(f"{'='*60}")

        for text_id, expected, text in sentences:
            done += 1
            print(f"\n[{done}/{total}] {model_key.upper()} | {text_id} | {text[:50]}...")
            t0 = time.time()

            try:
                row = analyze_text(
                    text_id, text, expected, model_key, commit_hash,
                    n_steps=args.n_steps,
                    bar_threshold=args.bar_threshold,
                    sig_alpha=args.sig_alpha,
                )
                rows.append(row)
                elapsed = time.time() - t0

                nb_str = (f"NB={row['Prob. classificador Notebook']}"
                          if row['Prob. classificador Notebook'] != "" else "NB=n/a")
                print(f"    -> {nb_str}  GUS={row['Prob. classificador GUS-NET']:.3f} "
                      f"({row['Bias Level (token-level)']})  "
                      f"BAR={row['Max bias ratio']:.3f}  "
                      f"IG_rho={row['Mean IG rho']:.3f}  "
                      f"Abl={row['Max Ablation Delta']:.6f}  "
                      f"LRP={row['rho(LRP, IG)']:.3f}  "
                      f"[{elapsed:.1f}s]")
            except Exception as e:
                print(f"    !! ERROR: {e}")
                import traceback; traceback.print_exc()
                rows.append({
                    "Data": datetime.now().strftime("%Y-%m-%d"),
                    "Modelo": model_key.upper(),
                    "ID do texto": text_id,
                    "Texto colado": text[:80],
                    "Notas / observacoes": f"ERROR: {e}",
                    "Passou? (S/N)": "",
                })

    # Build DataFrame with new results
    df_new = pd.DataFrame(rows)

    # Load existing Registo from --old-runs if specified, else from input file
    df_old = None
    old_runs_path = args.old_runs if args.old_runs else (args.input if is_input_xlsx else None)

    if old_runs_path:
        try:
            df_old = pd.read_excel(old_runs_path, sheet_name="Registo")
            print(f"Loaded {len(df_old)} existing executions from {old_runs_path}")
        except Exception:
            pass  # No existing Registo sheet

    # Normalize column names to match (old Excel has slightly different names)
    if df_old is not None:
        col_mapping = {
            "Versão da ferramenta / commit": "Versao / commit",
            "Prob. classificador (0–1) Notebook  ": "Prob. classificador Notebook",
            "Prob. classificador (0–1)  GUS-NET - Thresholds: Low < 0.15 | Moderate < 0.40 | High ≥ 0.40": "Prob. classificador GUS-NET",
            "Etiqueta classificador GUS-NET ": "Etiqueta classificador GUS-NET",
            "Top heads (Attention×Bias) (ex.: L8-H3; L10-H7)": "Top heads (Attention-Bias)",
            "Screenshots guardados? (S/N)": "Screenshots guardados? (S/N)",
            "Notas / observações": "Notas / observacoes",
            "Acção de follow-up": "Accao de follow-up",
            "Execução #": "Execucao #",
        }
        df_old = df_old.rename(columns=col_mapping)

        # Add missing columns with empty values
        for col in df_new.columns:
            if col not in df_old.columns:
                df_old[col] = ""

        # Ensure column order matches df_new
        df_old = df_old[[c for c in df_new.columns if c in df_old.columns] +
                        [c for c in df_old.columns if c not in df_new.columns]]

        # Concatenate old + new
        df = pd.concat([df_old, df_new], ignore_index=True)
    else:
        df = df_new

    # Renumber executions
    if "Execucao #" not in df.columns:
        df.insert(0, "Execucao #", range(1, len(df) + 1))
    else:
        df["Execucao #"] = range(1, len(df) + 1)

    # Save — if input is an xlsx and --out was not explicitly set, write back
    # into the same file (preserving Textos and other sheets).
    out_path = args.out
    is_input_xlsx = args.input and args.input.lower().endswith(".xlsx")
    if is_input_xlsx and args.out == "test_plan_results.xlsx":
        out_path = args.input  # write back to the same file

    if out_path.endswith(".xlsx"):
        preserve = is_input_xlsx and os.path.abspath(out_path) == os.path.abspath(args.input)
        _save_excel(df, out_path, preserve_sheets=preserve)
    else:
        df.to_csv(out_path, index=False, encoding="utf-8-sig")

    print(f"\n{'='*60}")
    print(f"Results saved to: {out_path}  ({len(df)} rows)")
    print(f"{'='*60}")

    # Summary
    summary_cols = [
        "ID do texto", "Modelo", "Prob. classificador Notebook",
        "Etiqueta classificador Notebook",
        "Prob. classificador GUS-NET", "Bias Level (token-level)",
        "Max bias ratio", "Mean BAR", "Mean IG rho",
        "Max Ablation Delta", "Mean Jaccard", "Mean RBO (p=0.9)",
        "rho(LRP, IG)", "Mean rho(LRP, Attn)",
        "rho(Perturb, IG)", "Max Perturbation Impact", "Passou? (S/N)",
    ]
    available = [c for c in summary_cols if c in df.columns]
    print("\n" + df[available].to_string(index=False))


if __name__ == "__main__":
    main()
